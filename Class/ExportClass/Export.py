import openpyxl

class Export_Excel:
    def __init__(self, _wb, _keyword, _nbResult, _language, _domaine):
        self.wb = _wb
        self.keyword = _keyword
        self.nb_result = _nbResult
        self.language = _language
        self.domaine = _domaine
        self.url = "https://www.google.com/search?q="
        self.link()
        self.init()

    def link(self):
        list_keywords = self.keyword.split()

        for i in range(len(list_keywords)):
            if i != len(list_keywords)-1:
                self.url+= list_keywords[i] + "+"
            else:
               self.url+= list_keywords[i] 

        self.url+= "&num=" + str(self.nb_result)

    def init(self):
        sheet = self.wb.active

        #case Web Scraper
        sheet.merge_cells('A1:C1')
        A1 = sheet.cell(row = 1, column = 1)
        A1.value = "WEB SCRAPER"

        #case de parametrage
        #mot clé
        sheet.merge_cells('A3:B3')
        A3 = sheet.cell(row = 3, column = 1)
        A3.value = "Mot clé : "
        sheet.merge_cells('C3:H3')
        C3 = sheet.cell(row = 3, column = 3)
        C3.value = self.keyword

        #nb resulat 
        sheet.merge_cells('A4:B4')
        A4 = sheet.cell(row = 4, column = 1)
        A4.value = "Nombre de résultat :"
        sheet.merge_cells('C4:H4')
        C4 = sheet.cell(row = 4, column = 3)
        C4.value = self.nb_result

        #langage 
        sheet.merge_cells('A5:B5')
        A5 = sheet.cell(row = 5, column = 1)
        A5.value = "Langage"
        sheet.merge_cells('C5:H5')
        C5 = sheet.cell(row = 5, column = 3)
        C5.value = self.language

        #domaine 
        sheet.merge_cells('A6:B6')
        A6 = sheet.cell(row = 6, column = 1)
        A6.value = "Domaine"
        sheet.merge_cells('C6:H6')
        C6 = sheet.cell(row = 6, column = 3)
        C6.value = self.domaine

        #url
        sheet.merge_cells('A7:B7')
        A7 = sheet.cell(row = 7, column = 1)
        A7.value = "URL"
        sheet.merge_cells('C7:H7')
        C7 = sheet.cell(row = 7, column = 3)
        C7.value = self.url


        #case de liste de résultats 
        #url 
        sheet.merge_cells('A10:C11')
        A10 = sheet.cell(row = 10, column = 1)
        A10.value = "URL"

        #titre 
        sheet.merge_cells('D10:F11')
        C10 = sheet.cell(row = 10, column = 4)
        C10.value = "Titre"

        #description
        sheet.merge_cells('G10:J11')
        G10 = sheet.cell(row = 10, column = 7)
        G10.value = "Description"

    def save_data(self, list_result):
        sheet = self.wb.active

        dec = 12
        for i in range(len(list_result)):
            #url 
            sheet.merge_cells("A"+str(dec)+":C"+str(dec+1))
            cell_url = sheet.cell(row = dec, column = 1)
            cell_url.value = list_result[i][0]

            #titre
            sheet.merge_cells("D"+str(dec)+":F"+str(dec+1))
            cell_titre = sheet.cell(row = dec, column = 4)
            cell_titre.value = list_result[i][1]

            #description
            sheet.merge_cells("G"+str(dec)+":J"+str(dec+1))
            cell_descr = sheet.cell(row = dec, column = 7)
            cell_descr.value = list_result[i][2]

            dec+=2

    